# export_results.py — تصدير النتائج إلى ملف Excel

import json
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from db import Database
from search import _score_to_label


# ─────────────────────────────────────────────────
# التصدير الرئيسي
# ─────────────────────────────────────────────────

def export_all_records(db: Database, output_path: str | None = None) -> str:
    """
    تصدير كل السجلات النشطة إلى ملف Excel.
    يُرجع مسار الملف المُصدَّر.
    """
    records = db.list_records(include_deleted=False, limit=100_000)
    df = _records_to_df(records)

    output_path = output_path or _default_output_path('all_records')
    df.to_excel(output_path, index=False, sheet_name='كل السجلات')
    _style_excel(output_path)
    return output_path


def export_match_results(match_results: list[dict],
                         output_path: str | None = None) -> str:
    """
    تصدير نتائج المطابقة إلى Excel مع شيتات متعددة:
    - موجود
    - محتمل موجود
    - يحتاج مراجعة
    - غير موجود
    - تقرير إحصائي
    """
    output_path = output_path or _default_output_path('match_results')

    # تصنيف النتائج
    found      = [r for r in match_results if r['result'] == 'موجود']
    probable   = [r for r in match_results if 'محتمل' in r['result']]
    review     = [r for r in match_results if 'مراجعة' in r['result']]
    not_found  = [r for r in match_results if 'غير موجود' in r['result']]

    def to_flat_df(items):
        rows = []
        for item in items:
            best = item['matches'][0] if item['matches'] else {}
            rows.append({
                'الاسم الأصلي':      item.get('query', ''),
                'الاسم المنظف':      item.get('query_clean', ''),
                'النتيجة':           item.get('result', ''),
                'أعلى نسبة تشابه':  item.get('best_score', 0),
                'أفضل مطابقة':       best.get('original_name', ''),
                'الشيت المطابق':     best.get('source_sheet', ''),
                'الملف المطابق':     best.get('source_file', ''),
                'الصف الأصلي':       best.get('original_row', ''),
                'صف المقارنة':       item.get('source_row', ''),
            })
        return pd.DataFrame(rows)

    # تقرير إحصائي
    total = len(match_results)
    stats_df = pd.DataFrame([
        {'البيان': 'إجمالي المقارنة',                'العدد': total, 'النسبة': '100%'},
        {'البيان': 'موجود (100%)',                    'العدد': len(found),     'النسبة': _pct(len(found), total)},
        {'البيان': 'محتمل موجود (95-99%)',           'العدد': len(probable),  'النسبة': _pct(len(probable), total)},
        {'البيان': 'يحتاج مراجعة (90-94%)',          'العدد': len(review),    'النسبة': _pct(len(review), total)},
        {'البيان': 'غير موجود (أقل من 90%)',         'العدد': len(not_found), 'النسبة': _pct(len(not_found), total)},
    ])

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        to_flat_df(found).to_excel(writer,     sheet_name='موجود',           index=False)
        to_flat_df(probable).to_excel(writer,  sheet_name='محتمل موجود',     index=False)
        to_flat_df(review).to_excel(writer,    sheet_name='يحتاج مراجعة',   index=False)
        to_flat_df(not_found).to_excel(writer, sheet_name='غير موجود',       index=False)
        stats_df.to_excel(writer,              sheet_name='تقرير إحصائي',    index=False)

    _style_excel(output_path)
    return output_path


# ─────────────────────────────────────────────────
# دوال مساعدة
# ─────────────────────────────────────────────────

def _records_to_df(records: list[dict]) -> pd.DataFrame:
    """تحويل قائمة السجلات إلى DataFrame مناسب للتصدير"""
    rows = []
    for r in records:
        base = {
            'id':           r['id'],
            'source_file':  r['source_file'],
            'source_sheet': r['source_sheet'],
            'original_row': r['original_row'],
            'original_name': r['original_name'],
            'clean_name':   r['clean_name'],
            'created_at':   r['created_at'],
            'updated_at':   r['updated_at'],
            'is_deleted':   r['is_deleted'],
        }
        # دمج بيانات الصف الأصلية
        row_data = r.get('row_data') or {}
        if isinstance(row_data, str):
            try:
                row_data = json.loads(row_data)
            except Exception:
                row_data = {}
        base.update(row_data)
        rows.append(base)

    return pd.DataFrame(rows)


def _default_output_path(prefix: str) -> str:
    """مسار افتراضي للتصدير بجانب قاعدة البيانات"""
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_dir = Path(__file__).parent / 'data'
    out_dir.mkdir(parents=True, exist_ok=True)
    return str(out_dir / f'{prefix}_{ts}.xlsx')


def _style_excel(file_path: str):
    """تنسيق أساسي لملف Excel: عناوين عريضة + ضبط عرض الأعمدة"""
    try:
        wb = load_workbook(file_path)
        header_font  = Font(bold=True, color='FFFFFF')
        header_fill  = PatternFill('solid', fgColor='1F4E79')
        header_align = Alignment(horizontal='center', vertical='center')

        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = header_align

            for col in ws.columns:
                max_len = max(
                    (len(str(c.value)) for c in col if c.value), default=8
                )
                ws.column_dimensions[get_column_letter(col[0].column)].width = \
                    min(max_len + 4, 40)

        wb.save(file_path)
    except Exception:
        pass  # التنسيق اختياري ولا يجب أن يوقف البرنامج


def _pct(part: int, total: int) -> str:
    if total == 0:
        return '0%'
    return f"{part / total * 100:.1f}%"
