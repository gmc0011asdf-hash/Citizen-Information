# match_excel.py — مطابقة ملف Excel مع قاعدة البيانات
# التشغيل: python match_excel.py

import sys
import os
import json
import time

# ضمان دعم UTF-8 على Windows
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import tkinter as tk
from tkinter import filedialog, messagebox
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz, process as fuzz_process

from db import Database
from cleaner import clean_arabic_name, detect_name_column, prepare_columns

# ─────────────────────────────────────────────────
# الثوابت
# ─────────────────────────────────────────────────

EXPORTS_DIR = Path(__file__).parent / "data" / "exports"

SCORE_EXACT      = 100
SCORE_HIGH_MIN   = 95
SCORE_REVIEW_MIN = 90

LABEL_FOUND    = "موجود"
LABEL_PROBABLE = "محتمل موجود بدرجة عالية"
LABEL_REVIEW   = "يحتاج مراجعة"
LABEL_MISSING  = "غير موجود"


# ─────────────────────────────────────────────────
# تحميل قاعدة البيانات في الذاكرة
# ─────────────────────────────────────────────────

def load_db_index(db: Database, source_file: str = None) -> tuple[dict, list, list]:
    """
    تحميل كل أسماء قاعدة البيانات إلى الذاكرة مرة واحدة.
    يمكن تصفية السجلات حسب الملف المصدر لتجنب تحميل ملف المقارنة نفسه.

    يُرجع:
        exact_index  : {clean_name -> record_dict}  للمطابقة المباشرة O(1)
        clean_names  : [clean_name, ...]             لـ rapidfuzz
        records_list : [record_dict, ...]            بنفس ترتيب clean_names
    """
    print("  [*] تحميل قاعدة البيانات في الذاكرة...", end=' ', flush=True)
    t0 = time.time()

    query = """SELECT id, original_name, clean_name, source_file,
                      source_sheet, original_row
               FROM records
               WHERE is_deleted = 0
               AND clean_name IS NOT NULL
               AND clean_name != ''"""
    params = []
    if source_file:
        query += " AND source_file = ?"
        params.append(source_file)
    query += " ORDER BY id"

    with db._get_connection() as conn:
        rows = conn.execute(query, params).fetchall()

    records_list = [dict(r) for r in rows]
    clean_names  = [r['clean_name'] for r in records_list]
    exact_index  = {r['clean_name']: r for r in records_list}

    elapsed = time.time() - t0
    print(f"تم — {len(records_list):,} سجل في {elapsed:.1f}s")
    return exact_index, clean_names, records_list


# ─────────────────────────────────────────────────
# منطق المطابقة
# ─────────────────────────────────────────────────

def match_name(query_clean: str,
               exact_index: dict,
               clean_names: list,
               records_list: list) -> dict:
    """
    مطابقة اسم واحد (منظف مسبقًا) ضد الفهرس.

    المرحلة 1: Exact Match بواسطة dict — O(1)
    المرحلة 2: Fuzzy Match بواسطة rapidfuzz — فقط إذا لم تُوجد مطابقة مباشرة
    """
    if not query_clean:
        return _no_match(0, '')

    # ── المرحلة 1: Exact Match ──
    if query_clean in exact_index:
        rec = exact_index[query_clean]
        return {
            'match_type':       'exact',
            'similarity_score': 100.0,
            'label':            LABEL_FOUND,
            'matched_name':     rec['original_name'],
            'matched_file':     rec['source_file'],
            'matched_sheet':    rec['source_sheet'],
            'matched_row':      rec['original_row'],
            'record_id':        rec['id'],
        }

    # ── المرحلة 2: Fuzzy Match ──
    result = fuzz_process.extractOne(
        query_clean,
        clean_names,
        scorer=fuzz.token_sort_ratio,
        score_cutoff=SCORE_REVIEW_MIN - 1,   # لا نحسب ما هو أقل من 89
    )

    if result is None:
        return _no_match(0, 'fuzzy')

    best_name, best_score, best_idx = result
    rec = records_list[best_idx]

    return {
        'match_type':       'fuzzy',
        'similarity_score': float(best_score),
        'label':            _score_label(best_score),
        'matched_name':     rec['original_name'],
        'matched_file':     rec['source_file'],
        'matched_sheet':    rec['source_sheet'],
        'matched_row':      rec['original_row'],
        'record_id':        rec['id'],
    }


def _no_match(score: float, match_type: str) -> dict:
    return {
        'match_type':       match_type,
        'similarity_score': score,
        'label':            LABEL_MISSING,
        'matched_name':     '',
        'matched_file':     '',
        'matched_sheet':    '',
        'matched_row':      '',
        'record_id':        None,
    }


def _score_label(score: float) -> str:
    if score >= SCORE_EXACT:
        return LABEL_FOUND
    elif score >= SCORE_HIGH_MIN:
        return LABEL_PROBABLE
    elif score >= SCORE_REVIEW_MIN:
        return LABEL_REVIEW
    return LABEL_MISSING


# ─────────────────────────────────────────────────
# قراءة ملف المقارنة
# ─────────────────────────────────────────────────

def read_check_file(file_path: str) -> list[dict]:
    """
    قراءة ملف Excel للمقارنة وتحويله إلى قائمة من السجلات.

    يُرجع:
        [{
          'check_file': اسم الملف,
          'check_sheet': اسم الشيت,
          'check_row': رقم الصف في Excel,
          'original_name': الاسم الأصلي,
          'clean_name': الاسم المنظف,
        }, ...]
    """
    file_name = Path(file_path).name

    print(f"  [*] قراءة ملف المقارنة: {file_name}")

    ext = Path(file_path).suffix.lower()
    engine = 'openpyxl' if ext in ('.xlsx', '.xlsm', '.xlsb') else None

    sheets = pd.read_excel(
        file_path,
        sheet_name=None,
        dtype=str,
        header=0,
        engine=engine,
    )

    all_entries = []

    for sheet_name, df in sheets.items():
        if df.empty:
            continue

        # تنظيف أسماء الأعمدة
        clean_cols, _ = prepare_columns(df.columns.tolist())
        df.columns = clean_cols
        df = df.dropna(how='all')

        # الكشف عن عمود الاسم
        name_col = detect_name_column(clean_cols)

        if not name_col:
            print(f"    [!] الشيت [{sheet_name}]: لم يُعثر على عمود اسم — سيُتخطى")
            continue

        found_in_sheet = 0
        for excel_idx, row in df.iterrows():
            raw_name = str(row.get(name_col, '') or '').strip()
            if not raw_name or raw_name.lower() == 'nan':
                continue

            all_entries.append({
                'check_file':    file_name,
                'check_sheet':   str(sheet_name),
                'check_row':     excel_idx + 2,   # +2 لأن Excel يبدأ من 1 والعنوان صف 1
                'original_name': raw_name,
                'clean_name':    clean_arabic_name(raw_name),
            })
            found_in_sheet += 1

        print(f"    شيت [{sheet_name}]: {found_in_sheet} اسم")

    return all_entries


# ─────────────────────────────────────────────────
# تنفيذ المطابقة الكاملة — نسخة محسّنة بـ Batch
# ─────────────────────────────────────────────────

def run_matching(entries: list[dict],
                 exact_index: dict,
                 clean_names: list,
                 records_list: list) -> list[dict]:
    """
    مطابقة محسّنة بمرحلتين:
      المرحلة 1 — Exact: يعالج كل القائمة بـ dict lookup فوري
      المرحلة 2 — Fuzzy Batch: يجمع كل الأسماء غير المطابقة ويشغّل
                  cdist مرة واحدة بكل المعالجات المتاحة (workers=-1)
                  أسرع بكثير من extractOne لكل اسم على حدة
    """
    import numpy as np

    total = len(entries)
    print(f"\n  [*] بدء المطابقة — {total:,} اسم...")
    t0 = time.time()

    # ── المرحلة 1: Exact Match لكل القائمة ──
    results     = [None] * total
    fuzzy_idxs  = []           # مؤشرات الأسماء التي لم تُطابَق مباشرةً

    exact_hits = 0
    for i, entry in enumerate(entries):
        q = entry['clean_name']
        if not q:
            results[i] = _no_match(0, '')
            continue

        if q in exact_index:
            rec = exact_index[q]
            results[i] = {
                'match_type':       'exact',
                'similarity_score': 100.0,
                'label':            LABEL_FOUND,
                'matched_name':     rec['original_name'],
                'matched_file':     rec['source_file'],
                'matched_sheet':    rec['source_sheet'],
                'matched_row':      rec['original_row'],
                'record_id':        rec['id'],
            }
            exact_hits += 1
        else:
            fuzzy_idxs.append(i)

    elapsed_exact = time.time() - t0
    print(f"      Exact: {exact_hits:,} مطابقة في {elapsed_exact:.2f}s")
    print(f"      Fuzzy: {len(fuzzy_idxs):,} اسم متبقٍّ للمطابقة التقريبية...",
          flush=True)

    # ── المرحلة 2: Fuzzy Batch ──
    if fuzzy_idxs:
        fuzzy_queries = [entries[i]['clean_name'] for i in fuzzy_idxs]

        t1 = time.time()

        # cdist: يحسب مصفوفة N×M كاملة بـ SIMD وكل المعالجات
        # score_cutoff يقطع الحسابات تحت الحد مبكرًا
        scores_matrix = fuzz_process.cdist(
            fuzzy_queries,
            clean_names,
            scorer=fuzz.token_sort_ratio,
            score_cutoff=SCORE_REVIEW_MIN - 1,
            workers=-1,         # استخدم كل أنوية المعالج
            dtype=np.uint8,
        )

        elapsed_cdist = time.time() - t1
        print(f"      cdist انتهى في {elapsed_cdist:.1f}s")

        fuzzy_hits = 0
        for list_pos, entry_idx in enumerate(fuzzy_idxs):
            row_scores = scores_matrix[list_pos]
            best_db_idx = int(row_scores.argmax())
            best_score  = float(row_scores[best_db_idx])

            if best_score < SCORE_REVIEW_MIN:
                results[entry_idx] = _no_match(best_score, 'fuzzy')
            else:
                rec = records_list[best_db_idx]
                label = _score_label(best_score)
                if label != LABEL_MISSING:
                    fuzzy_hits += 1
                results[entry_idx] = {
                    'match_type':       'fuzzy',
                    'similarity_score': best_score,
                    'label':            label,
                    'matched_name':     rec['original_name'],
                    'matched_file':     rec['source_file'],
                    'matched_sheet':    rec['source_sheet'],
                    'matched_row':      rec['original_row'],
                    'record_id':        rec['id'],
                }

        print(f"      Fuzzy hits: {fuzzy_hits:,}")

    elapsed = time.time() - t0
    print(f"  [✓] انتهت المطابقة الكاملة في {elapsed:.1f}s")

    # ── تجميع النتائج النهائية ──
    final = []
    for i, entry in enumerate(entries):
        match = results[i] or _no_match(0, '')
        final.append({
            # بيانات ملف المقارنة
            'اسم ملف المقارنة':       entry['check_file'],
            'شيت المقارنة':           entry['check_sheet'],
            'صف المقارنة':            entry['check_row'],
            'الاسم الأصلي':           entry['original_name'],
            'الاسم المنظف':           entry['clean_name'],
            # نتيجة المطابقة
            'النتيجة':                match['label'],
            'نسبة التشابه':           match['similarity_score'],
            'نوع المطابقة':           match['match_type'],
            # بيانات المطابق من قاعدة البيانات
            'أفضل اسم مطابق':         match['matched_name'],
            'ملف المطابق':            match['matched_file'],
            'شيت المطابق':            match['matched_sheet'],
            'صف المطابق في DB':       match['matched_row'],
        })

    return final


# ─────────────────────────────────────────────────
# تصدير النتائج
# ─────────────────────────────────────────────────

def export_results(results: list[dict], check_file_name: str) -> str:
    """
    تصدير النتائج إلى Excel متعدد الشيتات.
    يُرجع مسار ملف النتائج.
    """
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = str(EXPORTS_DIR / f"matching_result_{ts}.xlsx")

    df_all = pd.DataFrame(results)

    # تصنيف النتائج
    df_found    = df_all[df_all['النتيجة'] == LABEL_FOUND]
    df_probable = df_all[df_all['النتيجة'] == LABEL_PROBABLE]
    df_review   = df_all[df_all['النتيجة'] == LABEL_REVIEW]
    df_missing  = df_all[df_all['النتيجة'] == LABEL_MISSING]

    total = len(results)

    # التقرير الإحصائي
    stats = pd.DataFrame([
        {'البيان': 'إجمالي الأسماء المقارنة',         'العدد': total,             'النسبة': '100%'},
        {'البيان': f'موجود (100%)',                   'العدد': len(df_found),    'النسبة': _pct(len(df_found), total)},
        {'البيان': f'محتمل موجود (95-99%)',           'العدد': len(df_probable), 'النسبة': _pct(len(df_probable), total)},
        {'البيان': f'يحتاج مراجعة (90-94%)',          'العدد': len(df_review),   'النسبة': _pct(len(df_review), total)},
        {'البيان': f'غير موجود (أقل من 90%)',         'العدد': len(df_missing),  'النسبة': _pct(len(df_missing), total)},
        {'البيان': '─────────────────',               'العدد': '',               'النسبة': ''},
        {'البيان': 'ملف المقارنة',                   'العدد': check_file_name,  'النسبة': ''},
        {'البيان': 'تاريخ المطابقة',                 'العدد': datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'النسبة': ''},
    ])

    print(f"  [*] كتابة ملف النتائج...", end=' ', flush=True)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_all.to_excel(writer,      sheet_name='كل النتائج',      index=False)
        df_found.to_excel(writer,    sheet_name='موجود',            index=False)
        df_probable.to_excel(writer, sheet_name='محتمل موجود',      index=False)
        df_review.to_excel(writer,   sheet_name='يحتاج مراجعة',    index=False)
        df_missing.to_excel(writer,  sheet_name='غير موجود',        index=False)
        stats.to_excel(writer,       sheet_name='تقرير إحصائي',    index=False)

    _style_workbook(output_path)
    print("تم")
    return output_path


def _style_workbook(path: str):
    """تنسيق الملف — عناوين ملوّنة + ضبط عرض الأعمدة + تلوين حسب النتيجة"""
    try:
        wb = load_workbook(path)

        # ألوان حسب اسم الشيت
        SHEET_COLORS = {
            'موجود':         '1F7A4D',   # أخضر
            'محتمل موجود':  '2E75B6',   # أزرق
            'يحتاج مراجعة': 'C55A11',   # برتقالي
            'غير موجود':    'C00000',   # أحمر
            'كل النتائج':   '1F4E79',   # أزرق داكن
            'تقرير إحصائي': '404040',   # رمادي
        }

        header_font  = Font(bold=True, color='FFFFFF', size=11)
        header_align = Alignment(horizontal='center', vertical='center',
                                 wrap_text=True)

        for ws in wb.worksheets:
            color = SHEET_COLORS.get(ws.title, '1F4E79')
            fill  = PatternFill('solid', fgColor=color)

            # تنسيق صف العنوان
            ws.row_dimensions[1].height = 30
            for cell in ws[1]:
                cell.font      = header_font
                cell.fill      = fill
                cell.alignment = header_align

            # تلوين صفوف النتائج في شيت "كل النتائج"
            if ws.title == 'كل النتائج':
                _color_result_rows(ws)

            # ضبط عرض الأعمدة
            for col in ws.columns:
                max_len = max(
                    (len(str(c.value or '')) for c in col), default=8
                )
                ws.column_dimensions[
                    get_column_letter(col[0].column)
                ].width = min(max_len + 3, 45)

            # تجميد أول صف
            ws.freeze_panes = 'A2'

        wb.save(path)
    except Exception as e:
        print(f"    [تحذير] التنسيق: {e}")


def _color_result_rows(ws):
    """تلوين صفوف شيت 'كل النتائج' بحسب عمود النتيجة"""
    ROW_COLORS = {
        LABEL_FOUND:    'C6EFCE',   # أخضر فاتح
        LABEL_PROBABLE: 'DDEBF7',   # أزرق فاتح
        LABEL_REVIEW:   'FCE4D6',   # برتقالي فاتح
        LABEL_MISSING:  'FFDCE0',   # أحمر فاتح
    }

    # تحديد رقم عمود "النتيجة"
    result_col_idx = None
    for cell in ws[1]:
        if str(cell.value) == 'النتيجة':
            result_col_idx = cell.column
            break

    if result_col_idx is None:
        return

    for row in ws.iter_rows(min_row=2):
        result_cell = row[result_col_idx - 1]
        label = str(result_cell.value or '')
        bg = ROW_COLORS.get(label)
        if bg:
            fill = PatternFill('solid', fgColor=bg)
            for cell in row:
                cell.fill = fill


# ─────────────────────────────────────────────────
# واجهة المستخدم
# ─────────────────────────────────────────────────

def choose_file(title: str, filetypes=None) -> str | None:
    """فتح نافذة اختيار ملف"""
    if filetypes is None:
        filetypes = [
            ('ملفات Excel', '*.xlsx *.xls *.xlsm *.xlsb'),
            ('كل الملفات', '*.*'),
        ]
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    path = filedialog.askopenfilename(title=title, filetypes=filetypes)
    root.destroy()
    return path if path else None


def show_final_report(results: list[dict], output_path: str):
    """عرض التقرير النهائي في الطرفية ونافذة رسالة"""
    total     = len(results)
    found     = sum(1 for r in results if r['النتيجة'] == LABEL_FOUND)
    probable  = sum(1 for r in results if r['النتيجة'] == LABEL_PROBABLE)
    review    = sum(1 for r in results if r['النتيجة'] == LABEL_REVIEW)
    missing   = sum(1 for r in results if r['النتيجة'] == LABEL_MISSING)

    print(f"\n{'='*60}")
    print("  تقرير المطابقة النهائي")
    print(f"{'='*60}")
    print(f"  إجمالي الأسماء المقارنة : {total:,}")
    print(f"  موجود (100%)            : {found:,}  ({_pct(found, total)})")
    print(f"  محتمل موجود (95-99%)    : {probable:,}  ({_pct(probable, total)})")
    print(f"  يحتاج مراجعة (90-94%)  : {review:,}  ({_pct(review, total)})")
    print(f"  غير موجود (< 90%)       : {missing:,}  ({_pct(missing, total)})")
    print(f"{'='*60}")
    print(f"  ملف النتائج: {output_path}")
    print(f"{'='*60}\n")

    try:
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        msg = (
            f"اكتملت المطابقة!\n\n"
            f"إجمالي: {total:,} اسم\n"
            f"موجود: {found:,} ({_pct(found, total)})\n"
            f"محتمل موجود: {probable:,} ({_pct(probable, total)})\n"
            f"يحتاج مراجعة: {review:,} ({_pct(review, total)})\n"
            f"غير موجود: {missing:,} ({_pct(missing, total)})\n\n"
            f"ملف النتائج:\n{output_path}"
        )
        messagebox.showinfo('تقرير المطابقة', msg)
        root.destroy()
    except Exception:
        pass


# ─────────────────────────────────────────────────
# الدالة الرئيسية
# ─────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  نظام مطابقة Excel مع قاعدة البيانات")
    print("  Excel Matcher Against SQLite DB")
    print("=" * 60)

    # تهيئة قاعدة البيانات (قراءة فقط)
    db = Database()
    stats = db.get_stats()
    total_in_db = stats['records_with_name']

    if total_in_db == 0:
        print("\n  [!] قاعدة البيانات فارغة أو لا تحتوي أسماء.")
        print("      يرجى استيراد ملف Excel أولًا بواسطة: python main.py")
        return

    print(f"\n  قاعدة البيانات: {total_in_db:,} سجل بأسماء")
    print(f"  المسار: {stats['db_path']}\n")

    # ── اختيار ملف المقارنة ──
    print("  [*] يرجى اختيار ملف Excel للمقارنة...")
    check_file = choose_file('اختر ملف Excel للمقارنة')

    if not check_file:
        print("  [!] لم يتم اختيار أي ملف. إلغاء.")
        return

    print(f"  [*] ملف المقارنة: {check_file}\n")

    try:
        # ── تحميل DB في الذاكرة ──
        exact_index, clean_names, records_list = load_db_index(db)

        # ── قراءة ملف المقارنة ──
        entries = read_check_file(check_file)

        if not entries:
            print("  [!] لم يُعثر على أي اسم في ملف المقارنة.")
            return

        print(f"\n  [*] إجمالي الأسماء للمقارنة: {len(entries):,}")

        # ── تنفيذ المطابقة ──
        results = run_matching(entries, exact_index, clean_names, records_list)

        # ── التصدير ──
        output_path = export_results(results, Path(check_file).name)

        # ── التقرير ──
        show_final_report(results, output_path)

    except FileNotFoundError as e:
        print(f"\n  [!] الملف غير موجود: {e}")
    except ValueError as e:
        print(f"\n  [!] خطأ في البيانات: {e}")
    except Exception as e:
        import traceback
        print(f"\n  [!] خطأ غير متوقع: {e}")
        traceback.print_exc()


# ─────────────────────────────────────────────────
# مساعد
# ─────────────────────────────────────────────────

def _pct(part: int, total: int) -> str:
    if total == 0:
        return '0.0%'
    return f"{part / total * 100:.1f}%"


if __name__ == '__main__':
    main()
